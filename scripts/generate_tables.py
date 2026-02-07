#!/usr/bin/env python3
"""
Paper_01: 통합본 내 Table 1-7 + 부록 PRISMA -> 단일 Excel 파일 (탭 분류)
논문 본문에 내장된 마크다운 테이블을 추출하여 시트별 구성
"""

import re
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter

# --- Configuration ---
BASE_DIR = Path(__file__).parent
PAPER_DIR = BASE_DIR.parent.parent  # Paper_01_Context_Dependent/
SOURCE_FILE = PAPER_DIR / "PAPER_01_통합본_대한건축학회.md"
OUTPUT_FILE = BASE_DIR / "Paper_01_Tables_All.xlsx"

# --- Styles ---
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=10)
CAPTION_FILL = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
CAPTION_FONT = Font(name="Arial", bold=True, size=10)
NOTE_FONT = Font(name="Arial", italic=True, size=9, color="555555")
CELL_FONT = Font(name="Arial", size=10)
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)
WRAP_ALIGNMENT = Alignment(wrap_text=True, vertical="center")


def extract_tables_from_paper(filepath: Path):
    """Extract Table 1-7 and Appendix from the integrated paper.

    Strategy:
    - A table block starts with a "Table N." caption line (not inside a code block)
    - Then markdown table rows follow (|...|)
    - A "Note:" line may follow the table
    - The block ends when non-table content appears
    """
    text = filepath.read_text(encoding="utf-8")
    lines = text.split("\n")

    tables = []  # list of (table_id, caption, header_row, data_rows, note)
    i = 0
    in_code_block = False

    while i < len(lines):
        line = lines[i].strip()

        # Track code blocks
        if line.startswith("```"):
            in_code_block = not in_code_block
            i += 1
            continue

        if in_code_block:
            i += 1
            continue

        # Detect table caption: "Table N." at line start (not inside |)
        caption_match = re.match(r'^(Table \d+)\.\s+(.+)', line)
        # Also detect appendix table header
        appendix_match = re.match(r'^\|\s*섹션\s*\|', line) if not caption_match else None

        if caption_match:
            table_id = caption_match.group(1)
            caption = caption_match.group(2)
            i += 1

            # Skip blank lines until table starts
            while i < len(lines) and not lines[i].strip().startswith("|"):
                i += 1

            # Collect table rows
            table_lines = []
            while i < len(lines) and lines[i].strip().startswith("|"):
                row = lines[i].strip()
                # Skip separator rows: entire row is only |, -, :, whitespace
                if not (set(row) <= set("|-: ")):
                    table_lines.append(row)
                i += 1

            # Check for Note line
            note = ""
            while i < len(lines) and lines[i].strip() == "":
                i += 1
            if i < len(lines) and lines[i].strip().startswith("Note:"):
                note = lines[i].strip()
                i += 1

            if table_lines:
                header = _parse_row(table_lines[0])
                data = [_parse_row(r) for r in table_lines[1:]]
                tables.append((table_id, caption, header, data, note))
        elif appendix_match:
            # PRISMA checklist table in appendix
            table_lines = []
            while i < len(lines) and lines[i].strip().startswith("|"):
                row = lines[i].strip()
                if not (set(row) <= set("|-: ")):
                    table_lines.append(row)
                i += 1
            if table_lines:
                header = _parse_row(table_lines[0])
                data = [_parse_row(r) for r in table_lines[1:]]
                tables.append(("Appendix A", "PRISMA 2020 Checklist", header, data, ""))
        else:
            i += 1

    return tables


def _parse_row(line):
    cells = line.strip("|").split("|")
    return [c.strip() for c in cells]


def write_sheet(ws, caption, header, data_rows, note):
    """Write a single table to a worksheet."""
    row_idx = 1
    num_cols = len(header)

    # Caption row (merged)
    ws.cell(row=row_idx, column=1, value=caption)
    for c in range(1, num_cols + 1):
        cell = ws.cell(row=row_idx, column=c)
        cell.fill = CAPTION_FILL
        cell.font = CAPTION_FONT
        cell.border = THIN_BORDER
    if num_cols > 1:
        ws.merge_cells(
            start_row=row_idx, start_column=1,
            end_row=row_idx, end_column=num_cols,
        )
    row_idx += 1

    # Header row
    for c_idx, val in enumerate(header, start=1):
        cell = ws.cell(row=row_idx, column=c_idx, value=val)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.border = THIN_BORDER
        cell.alignment = WRAP_ALIGNMENT
    row_idx += 1

    # Data rows
    for data_row in data_rows:
        for c_idx, val in enumerate(data_row, start=1):
            cell = ws.cell(row=row_idx, column=c_idx, value=val)
            cell.font = CELL_FONT
            cell.border = THIN_BORDER
            cell.alignment = WRAP_ALIGNMENT
        row_idx += 1

    # Note row
    if note:
        row_idx += 1
        ws.cell(row=row_idx, column=1, value=note)
        for c in range(1, num_cols + 1):
            cell = ws.cell(row=row_idx, column=c)
            cell.font = NOTE_FONT
        if num_cols > 1:
            ws.merge_cells(
                start_row=row_idx, start_column=1,
                end_row=row_idx, end_column=num_cols,
            )

    # Auto-fit column widths
    for col in range(1, num_cols + 1):
        max_len = 0
        for row in ws.iter_rows(min_col=col, max_col=col, min_row=1, max_row=ws.max_row):
            for cell in row:
                if cell.value:
                    val_str = str(cell.value)
                    char_len = sum(2 if ord(c) > 127 else 1 for c in val_str)
                    if char_len > max_len:
                        max_len = char_len
        col_letter = get_column_letter(col)
        ws.column_dimensions[col_letter].width = min(max(max_len + 2, 10), 60)

    # Freeze header row
    ws.freeze_panes = "A3"


# Sheet name mapping
SHEET_NAMES = {
    "Table 1": "T1_연구특성",
    "Table 2": "T2_방향성분석",
    "Table 3": "T3_효과크기",
    "Table 4": "T4_이질성검정",
    "Table 5": "T5_하위그룹",
    "Table 6": "T6_메타회귀",
    "Table 7": "T7_주요발견",
    "Appendix A": "AppA_PRISMA",
}


def main():
    if not SOURCE_FILE.exists():
        print(f"[ERROR] Source file not found: {SOURCE_FILE}")
        return

    tables = extract_tables_from_paper(SOURCE_FILE)
    if not tables:
        print("[ERROR] No tables found in the paper.")
        return

    wb = Workbook()
    wb.remove(wb.active)

    for table_id, caption, header, data, note in tables:
        sheet_name = SHEET_NAMES.get(table_id, table_id)
        ws = wb.create_sheet(title=sheet_name)
        write_sheet(ws, caption, header, data, note)
        print(f"[OK] {sheet_name}: {len(data)} data rows x {len(header)} cols")

    wb.save(OUTPUT_FILE)
    print(f"\nSaved: {OUTPUT_FILE}")
    print(f"Total sheets: {len(wb.sheetnames)}")


if __name__ == "__main__":
    main()
